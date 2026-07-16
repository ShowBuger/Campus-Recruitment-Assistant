"""Excel template, export and import helpers for total-table records."""
from __future__ import annotations

import re
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel


HEADERS = [
    "公司名称", "秋招岗位", "城市", "批次", "嵌入式方向", "公司/行业类型",
    "优先级", "备注", "岗位JD", "投递链接", "投递截止时间", "投递时间",
    "机考时间", "一面", "二面", "三面", "保温", "结果", "进展",
]
DATE_HEADERS = {
    "投递截止时间", "投递时间", "机考时间", "一面", "二面", "三面", "保温", "结果",
}
LIST_HEADERS = {"嵌入式方向", "进展"}
VALID_BATCHES = {"秋招", "提前批"}
VALID_PROGRESS = {"未投递", "已投递", "机考", "面试", "OC", "已挂", "放弃"}
VALID_PRIORITIES = {"⭐⭐⭐⭐⭐", "⭐⭐⭐⭐", "⭐⭐⭐", "⭐⭐", "⭐"}
MAX_IMPORT_ROWS = 2000
MAX_FILE_BYTES = 10 * 1024 * 1024
CHINA_TZ = timezone(timedelta(hours=8))

_WIDTHS = [20, 26, 12, 12, 24, 20, 14, 28, 38, 32, 16, 16, 16, 16, 16, 16, 16, 16, 18]
_DATE_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日")


class ImportValidationError(ValueError):
    pass


def _safe_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    # Prevent spreadsheet formula injection in user-controlled exported values.
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


def _display_date(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)):
        try:
            return datetime.fromtimestamp(value / 1000, CHINA_TZ).strftime("%Y-%m-%d")
        except (OverflowError, OSError, ValueError):
            return str(value)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)[:10]


def _style_data_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="2563EB")
    thin = Side(style="thin", color="D7E0EE")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(1, ws.max_row)}"
    for index, width in enumerate(_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _add_instructions(workbook: Workbook) -> None:
    ws = workbook.create_sheet("填写说明")
    rows = [
        ("字段", "填写说明"),
        ("公司名称", "必填；空白行会被忽略"),
        ("秋招岗位", "选填，可以留空"),
        ("批次", "可填：秋招、提前批；留空默认为秋招"),
        ("进展", "可填：未投递、已投递、机考、面试、OC、已挂、放弃；留空默认为未投递"),
        ("嵌入式方向", "多个方向用中文顿号、逗号、分号或换行分隔"),
        ("嵌入式方向、公司/行业类型", "均为选填项，可以留空"),
        ("日期字段", "推荐使用 YYYY-MM-DD，例如 2026-08-15；也支持 Excel 日期单元格"),
        ("优先级", "可填 1 至 5 个星号，例如 ⭐⭐⭐"),
        ("导入限制", f"仅支持 .xlsx；单次最多 {MAX_IMPORT_ROWS} 条记录、文件最大 10 MB"),
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 72
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.font = Font(color="FFFFFF", bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_template() -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "总表导入模板"
    ws.append(HEADERS)
    _style_data_sheet(ws)
    _add_instructions(workbook)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_export(records: list[dict]) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "总表信息"
    ws.append(HEADERS)
    for record in records:
        fields = record.get("fields") or {}
        row = []
        for header in HEADERS:
            value = fields.get(header)
            if header in DATE_HEADERS:
                row.append(_display_date(value))
            elif header == "公司/行业类型":
                row.append(_safe_text((value or [""])[0] if isinstance(value, list) else value))
            elif header in LIST_HEADERS:
                row.append(_safe_text("、".join(str(item) for item in (value or []))))
            elif header == "投递链接" and isinstance(value, dict):
                row.append(_safe_text(value.get("link") or value.get("text") or ""))
            else:
                row.append(_safe_text(value))
        ws.append(row)
    _style_data_sheet(ws)
    _add_instructions(workbook)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _split_list(value) -> list[str]:
    if value in (None, ""):
        return []
    parts = re.split(r"[、,，;；\n\r]+", str(value))
    return list(dict.fromkeys(part.strip() for part in parts if part.strip()))


def _parse_date(value, row_number: int, header: str) -> int | None:
    if value in (None, ""):
        return None
    parsed: date | None = None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            parsed = converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            parsed = None
    else:
        text = str(value).strip()
        for fmt in _DATE_FORMATS:
            try:
                parsed = datetime.strptime(text, fmt).date()
                break
            except ValueError:
                continue
    if parsed is None or not 1900 <= parsed.year <= 2100:
        raise ImportValidationError(f"第 {row_number} 行「{header}」日期无效，请使用 YYYY-MM-DD")
    return int(datetime.combine(parsed, time.min, CHINA_TZ).timestamp() * 1000)


def parse_import(content: bytes) -> list[dict]:
    if not content:
        raise ImportValidationError("上传的 Excel 文件为空")
    if len(content) > MAX_FILE_BYTES:
        raise ImportValidationError("Excel 文件不能超过 10 MB")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportValidationError("无法读取 Excel，请确认文件是有效的 .xlsx 文件") from exc
    try:
        ws = workbook.worksheets[0]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        normalized = [str(value).strip() if value is not None else "" for value in first_row]
        duplicates = sorted({item for item in normalized if item and normalized.count(item) > 1})
        if duplicates:
            raise ImportValidationError("表头存在重复字段：" + "、".join(duplicates))
        header_map = {header: index for index, header in enumerate(normalized) if header in HEADERS}
        missing = [header for header in ("公司名称",) if header not in header_map]
        if missing:
            raise ImportValidationError("缺少必填表头：" + "、".join(missing))

        records: list[dict] = []
        errors: list[str] = []
        for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(value not in (None, "") for value in values):
                continue
            if len(records) >= MAX_IMPORT_ROWS:
                raise ImportValidationError(f"单次最多导入 {MAX_IMPORT_ROWS} 条记录")
            get = lambda header: values[header_map[header]] if header in header_map and header_map[header] < len(values) else None
            company = str(get("公司名称") or "").strip()
            job = str(get("秋招岗位") or "").strip()
            batch = str(get("批次") or "秋招").strip()
            progress_values = _split_list(get("进展")) or ["未投递"]
            priority = str(get("优先级") or "").strip()
            row_errors = []
            if not company:
                row_errors.append("公司名称不能为空")
            if batch not in VALID_BATCHES:
                row_errors.append("批次只能是秋招或提前批")
            invalid_progress = [item for item in progress_values if item not in VALID_PROGRESS]
            if invalid_progress:
                row_errors.append("进展值无效：" + "、".join(invalid_progress))
            if priority and priority not in VALID_PRIORITIES:
                row_errors.append("优先级应为 1 至 5 个星号")
            if row_errors:
                errors.append(f"第 {row_number} 行：" + "；".join(row_errors))
                continue
            try:
                fields = {
                    "公司名称": company,
                    "秋招岗位": job,
                    "城市": str(get("城市") or "").strip(),
                    "批次": batch,
                    "嵌入式方向": _split_list(get("嵌入式方向")),
                    "公司/行业类型": str(get("公司/行业类型") or "").strip(),
                    "优先级": priority,
                    "备注": str(get("备注") or "").strip(),
                    "岗位JD": str(get("岗位JD") or "").strip(),
                    "投递链接": str(get("投递链接") or "").strip(),
                    "进展": progress_values,
                }
                for header in DATE_HEADERS:
                    fields[header] = _parse_date(get(header), row_number, header)
                records.append(fields)
            except ImportValidationError as exc:
                errors.append(str(exc))
        if errors:
            visible = errors[:20]
            suffix = f"；另有 {len(errors) - 20} 处错误" if len(errors) > 20 else ""
            raise ImportValidationError("\n".join(visible) + suffix)
        if not records:
            raise ImportValidationError("Excel 中没有可导入的数据行")
        return records
    finally:
        workbook.close()
